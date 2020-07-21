import openpyxl

from globalData import data


def get_import(file_path, table_content):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[workbook.sheetnames[0]]
    able_list, unable_list = [], []
    data_list = []
    row_start, col_start = 0, 0
    for i in range(1, worksheet.max_row+1):
        for j in range(1, worksheet.max_column+1):
            if worksheet.cell(row=i, column=j).value == "名称":
                row_start, col_start = i, j-1
    for row in list(worksheet.rows)[row_start:]:
        list_tmp = []
        for cell in list(row)[col_start:col_start+8]:
            list_tmp.append(cell.value)
        data_list.append(list_tmp)
    print(data_list)
    for line in data_list:
        if able_to_import(line):
            line_tmp = [str(item) if i == 0 else int(item) for i, item in enumerate(line)]
            total_point, recommend_point = cal_tmp_point(line_tmp)
            line_tmp.extend([total_point, recommend_point])
            if line_tmp in able_list:
                continue
            if already_in_table(line_tmp[0], table_content):
                if not able_to_merge(line_tmp, table_content):
                    unable_list.append(line_tmp)
            else:
                able_list.append(line_tmp)
        else:
            if len(line) == 8:
                can_output = True
                for item in line:
                    if item is None:
                        can_output = False
                        break
                if can_output:
                    line_tmp = line
                    line_tmp.extend([0, 0])
                    unable_list.append(line_tmp)
    print(able_list, unable_list)
    return able_list, unable_list


def not_number(s):
    try:
        int(s)
        return False
    except ValueError:
        return True
    except TypeError:
        return True


def already_in_table(name, table_content):
    for item in table_content:
        if name == item[0]:
            return True
    return False


def able_to_merge(line, table_content):
    for item in table_content:
        if line[0] == item[0]:
            for i, element in enumerate(line[1:]):
                if element != item[i+1]:
                    return False
            return True


def able_to_import(line):
    if len(line) != 8:
        return False
    for element in line[1:]:
        if not_number(element) or int(element) > 100 or int(element) < 0:
            return False
    return True


def cal_tmp_point(line):
    weight_total = [data.WEIGHT_TOTAL['画风'], data.WEIGHT_TOTAL['画质'], data.WEIGHT_TOTAL['音乐'],
                    data.WEIGHT_TOTAL['语音'], data.WEIGHT_TOTAL['设定'], data.WEIGHT_TOTAL['剧情'],
                    data.WEIGHT_TOTAL['角色']]
    weight_recommend = [data.WEIGHT_RECOMMEND['画风'], data.WEIGHT_RECOMMEND['画质'],
                        data.WEIGHT_RECOMMEND['音乐'], data.WEIGHT_RECOMMEND['语音'],
                        data.WEIGHT_RECOMMEND['设定'], data.WEIGHT_RECOMMEND['剧情'],
                        data.WEIGHT_RECOMMEND['角色']]
    point_list = [int(line[1]), int(line[2]), int(line[3]), int(line[4]), int(line[5]), int(line[6]),
                  int(line[7])]
    total_point = 0
    recommend_point = 0
    for i, point in enumerate(point_list):
        total_point += point * weight_total[i]
        recommend_point += point * weight_recommend[i]
    total_point = round(total_point, 2)
    recommend_point = round(recommend_point, 2)
    return total_point, recommend_point


def table_import(file_path, table_now):
    pass


def table_export(file_path, table_now, table_content):
    index = len(table_content)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = table_now
    table_title = ["排名", "名称", "画风", "画质", "音乐", "语音", "设定", "剧情", "角色", "总分", "推荐"]
    for i in range(0, len(table_title)):
        sheet.cell(row=1, column=i + 1, value=table_title[i])
    for i in range(0, index):
        sheet.cell(row=i + 2, column=1, value=str(i + 1))
        for j in range(0, len(table_content[i])):
            sheet.cell(row=i + 2, column=j + 2, value=str(table_content[i][j]))
    try:
        workbook.save(file_path)
        # print("xlsx格式表格写入数据成功！")
        return True
    except PermissionError:
        return False


def import_table_out(file_path):
    table_title = ["名称", "画风(0-100)", "画质(0-100)", "音乐(0-100)", "语音(0-100)", "设定(0-100)",
                   "剧情(0-100)", "角色(0-100)"]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i in range(0, len(table_title)):
        sheet.cell(row=1, column=i + 1, value=table_title[i])
    try:
        workbook.save(file_path)
        # print("xlsx格式表格写入数据成功！")
        return True
    except PermissionError:
        return False
